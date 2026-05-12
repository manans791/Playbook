"""
Nium Playbook — Global Payout Capability Explorer
Integrated: Scraper + Data Layer + Dashboard (single file)

Folders created automatically:
  data/              — app reads from here (FI_data.xlsx, Non_FI_data.xlsx)
  scraped_data/      — archive copies with date (FI_2026-04-17.xlsx, etc.)
"""

import streamlit as st
import pandas as pd
import os, csv, time, glob, shutil
from io import BytesIO, StringIO
from datetime import datetime
from pathlib import Path
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ─── Paths ───────────────────────────────────────────────────────────────────
BASE_DIR = Path(os.getcwd())
DATA_DIR = BASE_DIR / "data"
ARCHIVE_DIR = BASE_DIR / "scraped_data"
DATA_DIR.mkdir(exist_ok=True)
ARCHIVE_DIR.mkdir(exist_ok=True)

FI_PATH = DATA_DIR / "FI_data.xlsx"
NON_FI_PATH = DATA_DIR / "Non_FI_data.xlsx"

# ═══════════════════════════════════════════════════════════════════════════════
# SCRAPER MODULE
# ═══════════════════════════════════════════════════════════════════════════════

COUNTRIES = [
    "Aland Islands","Albania","Algeria","American Samoa","Andorra","Angola","Anguilla","Antarctica",
    "Antigua","Argentina","Armenia","Aruba","Australia","Austria","Azerbaijan","Bahamas","Bahrain",
    "Bailiwick of Guernsey","Bangladesh","Barbados","Belgium","Belize","Benin","Bermuda","Bhutan",
    "Bolivia","Bonaire, Sint Eustatius and Saba","Bosnia and Herzegovina","Botswana","Bouvet Island",
    "Brazil","British Virgin Islands","Brunei","Bulgaria","Burkina Faso","Burundi","Cambodia","Cameroon",
    "Canada","Cape Verde","Caribbean Netherlands","Cayman Islands","Chad","Chile","China",
    "Christmas Island","Cocos (Keeling) Islands","Colombia","Comoros","Cook Islands","Costa Rica",
    "Croatia (Hrvatska)","Curaçao","Cyprus","Czech Republic","Democratic Republic of Timor-Leste",
    "Denmark","Diego Garcia","Djibouti","Dominica","Dominican Republic","Ecuador","Egypt","El Salvador",
    "Equatorial Guinea","Eritrea","Estonia","Falkland Islands (Malvinas)","Fiji","Finland","France",
    "French Guiana","French Polynesia","French Southern Territories","Gabon","Gambia","Georgia",
    "Germany","Ghana","Gibraltar","Greece","Greenland","Grenada","Guadeloupe","Guam","Guatemala",
    "Guinea","Guyana","Heard and Mc Donald Islands","Honduras","Hong Kong","Hungary","Iceland","India",
    "Indonesia","Ireland","Isle of Man","Israel","Italy","Ivory Coast","Jamaica","Japan","Jersey",
    "Jordan","Kazakhstan","Kenya","Kiribati","Kosovo","Kuwait","Kyrgyzstan","Laos","Latvia","Lesotho",
    "Liberia","Liechtenstein","Lithuania","Luxembourg","Macao","Macedonia","Madagascar","Malawi",
    "Malaysia","Maldives","Malta","Martinique","Mauritania","Mauritius","Mayotte","Mexico",
    "Micronesia, Federated States of","Moldova","Monaco","Mongolia","Montenegro","Montserrat",
    "Morocco","Mozambique","Namibia","Nauru","Nepal","Netherlands","New Caledonia","New Zealand",
    "Nicaragua","Niger","Nigeria","Niue","Norfolk Island","Northern Mariana Islands","Norway","Oman",
    "Pakistan","Palau","Palestine","Panama","Papua New Guinea","Paraguay","Peru","Philippines",
    "Pitcairn","Poland","Portugal","Puerto Rico","Qatar","Republic of the Congo","Reunion","Romania",
    "Rwanda","Saint Kitts and Nevis","Saint Lucia","Saint Martin (French part)",
    "Saint Vincent and the Grenadines","Samoa","San Marino","Sao Tome and Principe","Saudi Arabia",
    "Senegal","Serbia","Seychelles","Sierra Leone","Singapore","Sint Maarten (Dutch part)","Slovakia",
    "Slovenia","Solomon Islands","South Africa","South Georgia South Sandwich Islands","South Korea",
    "Spain","Sri Lanka","Suriname","Svalbard and Jan Mayen Islands","Swaziland","Sweden","Switzerland",
    "Taiwan","Tajikistan","Tanzania","Thailand","Togo","Tokelau","Tonga","Trinidad and Tobago",
    "Tunisia","Turkey","Turkmenistan","Turks and Caicos Islands","Tuvalu","Uganda","Ukraine",
    "United Arab Emirates","United Kingdom","United States Minor Outlying Islands",
    "United States of America","Uruguay","Uzbekistan","Vanuatu","Vatican City State","Vietnam",
    "Virgin Islands (U.S.)","Wallis and Futuna Islands","Western Sahara","Zambia",
]

def _format_country_url(name):
    return name.strip().lower().replace(" ", "-")

def _normalize_title(t):
    return t.replace("Bank Account (ACH) (BANK)", "Bank Account (ACH)").replace("Wallet (WALLET)", "Wallet").strip()

def _normalize_tat(t):
    return t.split("/")[0].strip()

def _extract_pills(container):
    pills = container.find_all("span", class_=lambda c: c and "rounded" in c)
    if pills:
        return ", ".join(p.get_text(strip=True) for p in pills)
    return container.get_text(strip=True)

def scrape_dataset(dataset_type, progress_bar, status_text):
    """
    Scrape playbook.nium.com for all countries.
    dataset_type: 'FI' or 'Non-FI'
    Returns: list of dicts (raw key-value rows)
    """
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.common.by import By
    from webdriver_manager.chrome import ChromeDriverManager
    from bs4 import BeautifulSoup

    url_suffix = "/institutions" if dataset_type == "FI" else ""

    options = webdriver.ChromeOptions()
    options.add_argument("--headless=new")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--log-level=3")
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

    raw_rows = []
    total = len(COUNTRIES)
    failed = []

    for idx, country_name in enumerate(COUNTRIES):
        pct = (idx + 1) / total
        progress_bar.progress(pct, text=f"Scraping {dataset_type}: {country_name} ({idx+1}/{total})")
        status_text.caption(f"⏳ {country_name}...")

        url_country = _format_country_url(country_name)
        url = f"https://playbook.nium.com/country/{url_country}{url_suffix}"

        try:
            driver.get(url)
            WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.CLASS_NAME, "payouts-details-table"))
            )
        except Exception:
            failed.append(country_name)
            continue

        soup = BeautifulSoup(driver.page_source, "html.parser")
        spans = soup.find_all("span", class_="payouts-details-table")

        for span in spans:
            try:
                raw_title = span.find(string=True, recursive=False)
                if not raw_title:
                    continue
                title = _normalize_title(raw_title.strip())

                currency, tat = "", ""
                small_tag = span.find("small")
                if small_tag:
                    lines = [l.strip() for l in small_tag.stripped_strings if l.strip()]
                    if lines:
                        currency = lines[0]
                    if len(lines) > 1:
                        tat = _normalize_tat(lines[1])

                modes_div = span.find("div", class_="hidden")
                modes = modes_div.get_text(strip=True) if modes_div else ""

                parent_h3 = span.find_parent("h3")
                if not parent_h3:
                    continue
                body_div = parent_h3.find_next("div", id=lambda x: x and "accordion-collapse-body" in x)
                if not body_div:
                    continue

                # Key-value blocks
                payout_blocks = body_div.find_all(
                    "div", class_="overflow-hidden bg-white border mb-5 font-normal payouts-details-table"
                )
                for block in payout_blocks:
                    rows = block.find_all("div", class_="bg-gray-50 px-4 py-5 sm:grid sm:grid-cols-3 sm:gap-4 sm:px-6")
                    for row in rows:
                        key_tag = row.find("dt")
                        value_tag = row.find("dd")
                        if not key_tag or not value_tag:
                            continue
                        key = key_tag.get_text(strip=True)
                        if key in ("Mandatory data requirements", "Supporting Documents"):
                            value = _extract_pills(value_tag)
                        else:
                            value = value_tag.get_text(strip=True)
                        raw_rows.append({
                            "Country": country_name, "Payment Mode": title,
                            "Currency": currency, "TAT": tat,
                            "Supported Modes": modes, "Key": key, "Value": value
                        })

                # Tables (transaction limits)
                for table in body_div.find_all("table"):
                    table_title = "Transaction limit per end-user"
                    parent_div = table.find_parent("div")
                    if parent_div:
                        dt_tag = parent_div.find_previous_sibling("dt") or parent_div.find("dt")
                        if not dt_tag:
                            outer_div = parent_div.find_parent("div")
                            if outer_div:
                                dt_tag = outer_div.find("dt")
                        if dt_tag:
                            table_title = dt_tag.get_text(strip=True)

                    headers = []
                    thead = table.find("thead")
                    if thead:
                        headers = [th.get_text(strip=True) for th in thead.find_all("th") if th.get_text(strip=True)]

                    tbody = table.find("tbody")
                    if tbody:
                        for tr in tbody.find_all("tr"):
                            cells = tr.find_all(["th", "td"])
                            if not cells:
                                continue
                            row_label = cells[0].get_text(strip=True)
                            for ci, cell in enumerate(cells[1:]):
                                if ci < len(headers):
                                    compound_key = f"{table_title} - {headers[ci]} - {row_label}"
                                    raw_rows.append({
                                        "Country": country_name, "Payment Mode": title,
                                        "Currency": currency, "TAT": tat,
                                        "Supported Modes": modes, "Key": compound_key,
                                        "Value": cell.get_text(strip=True)
                                    })
            except Exception:
                continue

    driver.quit()
    return raw_rows, failed


def transform_raw_to_wide(raw_rows):
    """
    Convert long-format key-value rows into wide-format DataFrame.
    Each unique (Country, Payment Mode, Currency, TAT, Supported Modes) gets one row
    with columns for each Key.
    """
    if not raw_rows:
        return pd.DataFrame()

    df_long = pd.DataFrame(raw_rows)

    # Group by corridor identity columns
    id_cols = ["Country", "Payment Mode", "Currency", "TAT", "Supported Modes"]

    # Pivot: each Key becomes a column
    # Handle duplicates by taking first value
    df_wide = df_long.pivot_table(
        index=id_cols,
        columns="Key",
        values="Value",
        aggfunc="first"
    ).reset_index()

    # Flatten column names
    df_wide.columns.name = None

    # Reorder: put identity columns first, then sort the rest
    other_cols = [c for c in df_wide.columns if c not in id_cols]
    
    # Preferred column order (match existing Excel structure)
    preferred_order = [
        "Supported Modes 1", "Supported Currencies", "Network Participant",
        "Channels", "Cutoff & delivery timing", "Mandatory data requirements",
        "Supporting Documents", "Beneficiary Statement Narrative", "Proof of Payment", "Notes",
    ]
    
    ordered = []
    for col in preferred_order:
        if col in other_cols:
            ordered.append(col)
            other_cols.remove(col)
    ordered += sorted(other_cols)  # Remaining columns alphabetically

    df_wide = df_wide[id_cols + ordered]

    return df_wide


def save_scraped_data(df, dataset_type):
    """Save to data/ (app source) and scraped_data/ (archive with date)"""
    today = datetime.now().strftime("%Y-%m-%d")
    
    # App source
    app_path = FI_PATH if dataset_type == "FI" else NON_FI_PATH
    df.to_excel(str(app_path), index=False)
    
    # Archive
    archive_path = ARCHIVE_DIR / f"{dataset_type}_{today}.xlsx"
    df.to_excel(str(archive_path), index=False)
    
    return str(app_path), str(archive_path)


def run_full_scrape(progress_bar, status_text):
    """Run scrape for both FI and Non-FI, transform, save."""
    results = {}
    all_failed = {}
    
    for ds_type in ["FI", "Non-FI"]:
        status_text.caption(f"🚀 Starting {ds_type} scrape...")
        raw_rows, failed = scrape_dataset(ds_type, progress_bar, status_text)
        all_failed[ds_type] = failed
        
        status_text.caption(f"🔄 Transforming {ds_type} data ({len(raw_rows)} raw rows)...")
        df = transform_raw_to_wide(raw_rows)
        
        if not df.empty:
            app_path, archive_path = save_scraped_data(df, ds_type)
            results[ds_type] = {
                "rows": len(df),
                "app_path": app_path,
                "archive_path": archive_path,
                "failed_countries": failed
            }
        else:
            results[ds_type] = {"rows": 0, "failed_countries": failed}
    
    return results


# ═══════════════════════════════════════════════════════════════════════════════
# DATA LAYER
# ═══════════════════════════════════════════════════════════════════════════════

def get_last_updated():
    """Get modification timestamp of data files"""
    if FI_PATH.exists():
        ts = os.path.getmtime(str(FI_PATH))
        return datetime.fromtimestamp(ts).strftime("%d %b %Y, %I:%M %p")
    return None

@st.cache_data(ttl=300)
def load_data_cached():
    """Load data from data/ folder. Falls back to any Excel in current dir."""
    fi_path = str(FI_PATH) if FI_PATH.exists() else None
    nfi_path = str(NON_FI_PATH) if NON_FI_PATH.exists() else None
    
    # Fallback: look for Excel files in current directory (backward compat)
    if not fi_path or not nfi_path:
        cd = os.getcwd()
        for f in os.listdir(cd):
            if f.endswith('.xlsx'):
                if 'FI' in f and 'Non' not in f and fi_path is None:
                    fi_path = f
                elif 'Non' in f and 'FI' in f and nfi_path is None:
                    nfi_path = f
    
    if not fi_path or not nfi_path:
        return None, None
    
    fi = pd.read_excel(fi_path)
    nfi = pd.read_excel(nfi_path)
    
    # Fix mixed types for Arrow serialization
    for d in [fi, nfi]:
        for c in d.columns:
            if 'Transaction limit' in c:
                d[c] = d[c].astype(str)
    
    fi['_t'] = 'FI'
    nfi['_t'] = 'Non-FI'
    return fi, nfi


# ═══════════════════════════════════════════════════════════════════════════════
# PAGE CONFIG & STYLING
# ═══════════════════════════════════════════════════════════════════════════════

st.set_page_config(page_title="Nium Playbook", page_icon="⚡", layout="wide", initial_sidebar_state="collapsed")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700;800&display=swap');

*, .main, .block-container, .stApp { font-family: 'Plus Jakarta Sans', sans-serif !important; }
.block-container { padding: 1rem 2rem 2rem 2rem; max-width: 100%; }
#MainMenu, footer, header { visibility: hidden; }
hr { margin: 0.5rem 0; opacity: 0.08; }

.n-hdr {
    display: flex; align-items: center; justify-content: space-between;
    padding: 1rem 1.5rem; background: #0A0A0A; border-radius: 14px;
    margin-bottom: 1rem; position: relative; overflow: hidden;
    border: 1px solid rgba(0,212,170,0.15);
}
.n-hdr::before { content:''; position:absolute; top:-60%; right:-8%; width:280px; height:280px; background:radial-gradient(circle,rgba(0,212,170,0.12) 0%,transparent 65%); pointer-events:none; }
.n-hdr::after { content:''; position:absolute; bottom:-40%; left:20%; width:200px; height:200px; background:radial-gradient(circle,rgba(0,229,191,0.06) 0%,transparent 65%); pointer-events:none; }
.n-logo { display:flex; align-items:center; gap:0.7rem; }
.n-icon { width:36px; height:36px; background:linear-gradient(135deg,#00D4AA,#00E5BF); border-radius:9px; display:flex; align-items:center; justify-content:center; font-size:1.1rem; font-weight:800; color:#0A0A0A; box-shadow:0 0 20px rgba(0,212,170,0.3); }
.n-hdr h1 { color:#fff; font-size:1.3rem; font-weight:700; margin:0; padding:0; border:none; letter-spacing:-0.02em; }
.n-hdr .n-sub { color:#00D4AA; font-size:0.68rem; font-weight:600; letter-spacing:0.08em; text-transform:uppercase; }
.n-hdr .n-updated { color:#64748b; font-size:0.65rem; font-weight:400; }

.stRadio > div { flex-direction:row !important; gap:0; }
.stRadio > div > label { background:#f7faf9 !important; padding:0.35rem 1.2rem !important; border:1px solid #d1e8e0 !important; font-weight:600 !important; font-size:0.78rem !important; color:#2d6a5a !important; margin:0 !important; }
.stRadio > div > label:first-child { border-radius:8px 0 0 8px !important; border-right:none !important; }
.stRadio > div > label:last-child { border-radius:0 8px 8px 0 !important; }
.stRadio > div > label[data-checked="true"] { background:#0A0A0A !important; color:#00D4AA !important; border-color:#0A0A0A !important; }

.n-stats { display:flex; gap:0.5rem; }
.n-sc { flex:1; background:#f7faf9; border:1px solid #d1e8e0; border-radius:10px; padding:0.55rem 0.9rem; border-left:3px solid #00D4AA; }
.n-sc .v { font-size:1.4rem; font-weight:800; color:#0A0A0A; line-height:1.2; letter-spacing:-0.03em; }
.n-sc .l { font-size:0.6rem; color:#00997A; font-weight:700; text-transform:uppercase; letter-spacing:0.07em; }

.n-fc { background:#fff; border:1px solid #e0efe8; border-radius:12px; padding:0.8rem 1rem; margin-bottom:0.6rem; border-top:3px solid #00D4AA; }
.n-fc-t { font-size:0.65rem; font-weight:700; color:#00997A; text-transform:uppercase; letter-spacing:0.08em; margin-bottom:0.3rem; }

.stSelectbox label, .stMultiSelect label { font-size:0.72rem !important; font-weight:600 !important; color:#2d6a5a !important; text-transform:uppercase !important; letter-spacing:0.04em !important; }
.stCheckbox label span { font-size:0.78rem !important; font-weight:500 !important; color:#1a1a1a !important; }
.stCheckbox { margin-bottom:-0.6rem; }
.stMultiSelect [data-baseweb="tag"] { background:#00D4AA !important; color:#0A0A0A !important; border-radius:6px !important; font-size:0.7rem !important; font-weight:700 !important; }
[data-testid="stMetric"] { display:none; }

.stDownloadButton > button { background:#0A0A0A !important; color:#00D4AA !important; border:1px solid rgba(0,212,170,0.3) !important; border-radius:8px !important; font-weight:700 !important; font-size:0.78rem !important; padding:0.45rem 1rem !important; }
.stDownloadButton > button:hover { background:#111 !important; border-color:#00D4AA !important; box-shadow:0 4px 15px rgba(0,212,170,0.2); transform:translateY(-1px); }
.stButton > button { border-radius:8px; font-weight:600; font-size:0.78rem; border:1px solid #d1e8e0; background:#0A0A0A !important; color:#00D4AA !important; border-color:rgba(0,212,170,0.3) !important; }
.stButton > button:hover { background:#111 !important; border-color:#00D4AA !important; box-shadow:0 3px 10px rgba(0,212,170,0.18); transform:translateY(-1px); }

[data-testid="stDataFrame"] { border:1px solid #e0efe8; border-radius:10px; overflow:hidden; }
.streamlit-expanderHeader { font-size:0.8rem !important; font-weight:600 !important; background:#f7faf9 !important; border-radius:8px !important; color:#2d6a5a !important; }
.n-bdg { display:inline-block; background:rgba(0,212,170,0.12); color:#00997A; padding:0.18rem 0.6rem; border-radius:20px; font-size:0.7rem; font-weight:700; border:1px solid rgba(0,212,170,0.25); }
.stAlert { border-radius:10px; font-size:0.82rem; }
.n-ft { text-align:center; color:#94a3b8; font-size:0.68rem; padding:1.2rem 0 0.5rem 0; margin-top:1.5rem; border-top:1px solid #e0efe8; }
.n-ft span { color:#00D4AA; font-weight:600; }
</style>
""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def cas(data, col, f):
    d = data.copy()
    for c, v in f.items():
        if c != col and v: d = d[d[c].isin(v)]
    return sorted(d[col].dropna().unique())

def fmt(v):
    if pd.isna(v) or str(v) in ('', 'nan', 'None'): return "—"
    return str(v)

def create_formatted_excel(data, selected_cols, dataset_type):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Nium Capabilities'
    blue_header = '4472C4'
    thin_border = Border(left=Side(style='thin',color='B4B4B4'), right=Side(style='thin',color='B4B4B4'), top=Side(style='thin',color='B4B4B4'), bottom=Side(style='thin',color='B4B4B4'))
    header_border = Border(left=Side(style='thin',color='FFFFFF'), right=Side(style='thin',color='FFFFFF'), top=Side(style='thin',color='FFFFFF'), bottom=Side(style='medium',color='2F5496'))
    ws.row_dimensions[1].height = 8
    ws.merge_cells('A2:E2')
    title_cell = ws['A2']
    title_cell.value = "Nium Payout Capability Matrix"
    title_cell.font = Font(name='Segoe UI Semibold', size=14, bold=True, color='1A1A1A')
    title_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 6
    headers = ['#'] + selected_cols
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        cell.font = Font(name='Segoe UI Semibold', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color=blue_header, end_color=blue_header, fill_type='solid')
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell.border = header_border
    ws.row_dimensions[4].height = 24
    last_col_letter = get_column_letter(len(headers))
    ws.auto_filter.ref = f'A4:{last_col_letter}4'
    export_df = data[selected_cols].reset_index(drop=True)
    for row_idx, (_, row) in enumerate(export_df.iterrows()):
        excel_row = row_idx + 5
        sn_cell = ws.cell(row=excel_row, column=1)
        sn_cell.value = row_idx + 1
        sn_cell.font = Font(name='Segoe UI Semilight', size=11, color='333333')
        sn_cell.alignment = Alignment(horizontal='right', vertical='top')
        sn_cell.border = thin_border
        for col_idx, col_name in enumerate(selected_cols):
            cell = ws.cell(row=excel_row, column=col_idx + 2)
            val = row[col_name]
            cell.value = None if (pd.isna(val) or str(val) in ('nan','None','')) else str(val)
            cell.font = Font(name='Segoe UI Semilight', size=11, color='333333')
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = thin_border
    ws.column_dimensions['A'].width = 5
    for col_idx, col_name in enumerate(selected_cols):
        col_letter = get_column_letter(col_idx + 2)
        max_len = len(col_name)
        for ri in range(5, min(ws.max_row + 1, 30)):
            cv = ws.cell(row=ri, column=col_idx + 2).value
            if cv: max_len = max(max_len, min(len(str(cv)), 45))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)
    ws.freeze_panes = 'B5'
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════════════════════════════
# UI — HEADER
# ═══════════════════════════════════════════════════════════════════════════════

last_updated = get_last_updated()
updated_html = f'<div class="n-updated">Last updated: {last_updated}</div>' if last_updated else '<div class="n-updated">No data yet — click 🔄 Refresh</div>'

st.markdown(f"""
<div class="n-hdr">
    <div class="n-logo">
        <div class="n-icon">N</div>
        <div>
            <h1>Nium Playbook</h1>
            <div class="n-sub">Global Payout Capability Explorer</div>
            {updated_html}
        </div>
    </div>
</div>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════════
# UI — REFRESH BUTTON
# ═══════════════════════════════════════════════════════════════════════════════

if st.button("🔄 Refresh Data (Scrape from Nium Playbook)", use_container_width=True, key="refresh"):
    st.warning("⚡ Starting full scrape — this takes ~25 minutes. Do not close the browser.")
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    try:
        results = run_full_scrape(progress_bar, status_text)
        
        progress_bar.progress(1.0, text="✅ Scrape complete!")
        
        summary = []
        for ds_type, info in results.items():
            summary.append(f"**{ds_type}:** {info['rows']} corridors scraped")
            if info.get('failed_countries'):
                summary.append(f"  ⚠️ {len(info['failed_countries'])} countries skipped")
        
        status_text.success("✅ Scrape complete!\n\n" + "\n".join(summary))
        
        # Clear cache so new data loads
        st.cache_data.clear()
        time.sleep(2)
        st.rerun()
        
    except Exception as e:
        st.error(f"❌ Scrape failed: {str(e)}")
        st.info("Make sure Chrome is installed and you have internet access.")

# ═══════════════════════════════════════════════════════════════════════════════
# UI — LOAD & DISPLAY DATA
# ═══════════════════════════════════════════════════════════════════════════════

fi_data, non_fi_data = load_data_cached()

if fi_data is None or non_fi_data is None:
    st.info("📭 No data files found. Click **🔄 Refresh Data** above to scrape from Nium Playbook, or place Excel files in the `data/` folder.")
    st.stop()

# ── Toggle + Stats ──
r1, r2 = st.columns([1, 5])
with r1:
    ds = st.radio("_", ["FI", "Non-FI"], horizontal=True, label_visibility="collapsed")
data = (fi_data if ds == "FI" else non_fi_data).copy().drop('_t', axis=1)
with r2:
    st.markdown(f"""
    <div class="n-stats">
        <div class="n-sc"><div class="l">Corridors</div><div class="v">{len(data):,}</div></div>
        <div class="n-sc"><div class="l">Countries</div><div class="v">{data['Country'].nunique()}</div></div>
        <div class="n-sc"><div class="l">Pay Modes</div><div class="v">{data['Payment Mode'].nunique()}</div></div>
        <div class="n-sc"><div class="l">Currencies</div><div class="v">{data['Currency'].nunique()}</div></div>
    </div>
    """, unsafe_allow_html=True)

# ── Filters ──
st.markdown('<div class="n-fc"><div class="n-fc-t">⚡ Filters — cascading</div>', unsafe_allow_html=True)

f1, f2, f3, f4 = st.columns(4)
ch = {}

with f1:
    country = st.selectbox("Country", ["All"] + sorted(data['Country'].dropna().unique()), key="c")
if country != "All": ch['Country'] = [country]

with f2:
    modes = st.multiselect("Payment Mode", cas(data, 'Payment Mode', ch), key="m", placeholder="All")
if modes: ch['Payment Mode'] = modes

with f3:
    currencies = st.multiselect("Currency", cas(data, 'Currency', ch), key="cr", placeholder="All")
if currencies: ch['Currency'] = currencies

with f4:
    at = sorted(cas(data, 'TAT', ch), key=lambda x: {'Realtime':0,'T0':1,'T1':2,'T2':3}.get(x, 9))
    tats = st.multiselect("TAT", at, key="t", placeholder="All")

c1, c2, c3, c4 = st.columns(4)
with c1: b2b = st.checkbox("B2B")
with c2: b2p = st.checkbox("B2P")
with c3: p2p = st.checkbox("P2P")
with c4: p2b = st.checkbox("P2B")
txn = [x for x, v in [("B2B",b2b),("B2P",b2p),("P2P",p2p),("P2B",p2b)] if v]

st.markdown('</div>', unsafe_allow_html=True)

# ── Apply Filters ──
df = data.copy()
if country != "All": df = df[df['Country'] == country]
if modes: df = df[df['Payment Mode'].isin(modes)]
if currencies: df = df[df['Currency'].isin(currencies)]
if tats: df = df[df['TAT'].isin(tats)]
if txn: df = df[df.apply(lambda r: all(t in str(r['Supported Modes']).split(', ') for t in txn), axis=1)]

# ── Results Bar ──
active = sum([country != "All", bool(modes), bool(currencies), bool(tats), bool(txn)])
rc1, rc2 = st.columns([5, 1.5])
with rc1:
    st.markdown(f"""
    <div style="display:flex;align-items:center;gap:0.6rem;padding:0.3rem 0;">
        <span style="font-size:1.05rem;font-weight:700;color:#0A0A0A;">{len(df):,} results</span>
        <span class="n-bdg">{active} filter{"s" if active != 1 else ""}</span>
    </div>
    """, unsafe_allow_html=True)
with rc2:
    if len(df) > 0:
        excel_buf = create_formatted_excel(df, df.columns.tolist(), ds)
        st.download_button(
            "⬇ Download Excel", excel_buf,
            f"Nium_Capability_Matrix_{ds}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

# ── Table ──
if len(df) > 0:
    st.dataframe(df, use_container_width=True, height=520, hide_index=True)
else:
    st.info("No corridors match. Adjust filters above.")

# ── Footer ──
st.markdown(f'<div class="n-ft">Powered by <span>Nium</span> · Playbook · {datetime.now().strftime("%B %Y")}</div>', unsafe_allow_html=True)